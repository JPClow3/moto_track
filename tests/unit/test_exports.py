from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import RequestFactory, TestCase, override_settings
from django.utils import timezone
from openpyxl import load_workbook

from apps.core import exports as core_exports
from apps.core.services.demo_bike import create_demo_motorcycle
from apps.documents.export import build_export as build_documents_export
from apps.documents.models import DocumentType, MotorcycleDocument
from apps.garage.models import Motorcycle
from apps.maintenance.export import _amount
from apps.maintenance.export import build_export as build_maintenance_export
from apps.maintenance.models import MaintenanceRecord, MaintenanceType
from apps.reminders.export import build_export as build_reminders_export
from apps.reminders.models import Reminder, TriggerType
from apps.tires.export import build_export as build_tires_export
from apps.tires.models import TirePosition, TirePressureRecord, TireRecord


class CoreExportHelpersTests(TestCase):
    def test_build_csv_bytes_includes_headers_and_safe_strings(self):
        payload = core_exports.build_csv_bytes(
            rows=[{"name": "Moto A", "notes": None}],
            columns=[("name", "Moto"), ("notes", "Notas")],
        ).decode("utf-8-sig")

        self.assertEqual(payload.splitlines(), ["Moto,Notas", "Moto A,"])

    def test_build_xlsx_bytes_writes_rows(self):
        payload = core_exports.build_xlsx_bytes(
            rows=[{"name": "Moto A", "km": 1234}],
            columns=[("name", "Moto"), ("km", "KM")],
        )

        workbook = load_workbook(BytesIO(payload))
        sheet = workbook.active

        self.assertEqual(sheet.title, "Export")
        self.assertEqual(sheet["A1"].value, "Moto")
        self.assertEqual(sheet["B1"].value, "KM")
        self.assertEqual(sheet["A2"].value, "Moto A")
        self.assertEqual(sheet["B2"].value, 1234)

    def test_build_xlsx_bytes_requires_openpyxl(self):
        with patch.object(core_exports, "_OpenpyxlWorkbook", None):
            with self.assertRaises(RuntimeError):
                core_exports.build_xlsx_bytes(rows=[], columns=[])

    def test_export_response_sets_attachment_headers(self):
        response = core_exports.export_response(
            content=b"abc",
            filename="teste.csv",
            content_type="text/csv; charset=utf-8",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content, b"abc")
        self.assertEqual(response["Content-Disposition"], 'attachment; filename="teste.csv"')

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="noreply@mototrack.local",
    )
    def test_maybe_email_export_respects_recipient(self):
        mail.outbox = []

        core_exports.maybe_email_export(
            to_email=None,
            subject="Sem envio",
            body="Nada",
            attachment_name="teste.csv",
            attachment_bytes=b"abc",
            attachment_content_type="text/csv",
        )
        self.assertEqual(len(mail.outbox), 0)

        core_exports.maybe_email_export(
            to_email="destinatario@example.com",
            subject="Com envio",
            body="Segue anexo",
            attachment_name="teste.csv",
            attachment_bytes=b"abc",
            attachment_content_type="text/csv",
        )

        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ["destinatario@example.com"])
        self.assertEqual(mail.outbox[0].attachments[0][0], "teste.csv")

    def test_parse_date_param_handles_blank_invalid_and_valid_values(self):
        self.assertIsNone(core_exports.parse_date_param(None))
        self.assertIsNone(core_exports.parse_date_param(""))
        self.assertIsNone(core_exports.parse_date_param("2026-99-99"))
        self.assertEqual(core_exports.parse_date_param("2026-05-03"), date(2026, 5, 3))

    def test_safe_next_url_only_accepts_local_hosts(self):
        request = RequestFactory().get("/blog/", HTTP_HOST="localhost")

        self.assertEqual(
            core_exports.safe_next_url(request=request, candidate="/garage/", fallback="/"),
            "/garage/",
        )
        self.assertEqual(
            core_exports.safe_next_url(request=request, candidate="https://evil.example/phish", fallback="/"),
            "/",
        )
        self.assertEqual(core_exports.safe_next_url(request=request, candidate=None, fallback="/"), "/")


@override_settings(
    EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
    DEFAULT_FROM_EMAIL="noreply@mototrack.local",
)
class ExportModulesTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username="export-user", email="export@example.com", password="pass12345")
        self.other_user = User.objects.create_user(
            username="export-other", email="export-other@example.com", password="pass12345"
        )
        self.motorcycle = Motorcycle.objects.create(
            owner=self.user,
            name="Moto Export",
            brand="Honda",
            model="CB 500",
            year=2024,
        )
        self.other_motorcycle = Motorcycle.objects.create(
            owner=self.other_user,
            name="Moto Outra",
            brand="Yamaha",
            model="MT-03",
            year=2023,
        )
        mail.outbox = []

    def test_documents_export_scopes_rows_and_sends_csv_email(self):
        MotorcycleDocument.objects.create(
            motorcycle=self.motorcycle,
            name="CRLV 2026",
            document_type=DocumentType.CRLV,
            valid_until=date(2026, 12, 31),
            notes="Digital",
            file=SimpleUploadedFile("crlv.pdf", b"pdf", content_type="application/pdf"),
        )
        MotorcycleDocument.objects.create(
            motorcycle=self.other_motorcycle,
            name="Seguro de outro usuario",
            document_type=DocumentType.INSURANCE,
            file=SimpleUploadedFile("seguro.pdf", b"pdf", content_type="application/pdf"),
        )

        response = build_documents_export(
            user=self.user,
            start=date(2026, 1, 1),
            end=date(2026, 12, 31),
            fmt="csv",
            email_to="docs@example.com",
        )
        payload = response.content.decode("utf-8-sig")

        self.assertIn("CRLV 2026", payload)
        self.assertNotIn("Seguro de outro usuario", payload)
        self.assertEqual(response["Content-Disposition"], 'attachment; filename="documentos.csv"')
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].attachments[0][0], "documentos.csv")

    def test_maintenance_export_filters_xlsx_and_formats_costs(self):
        MaintenanceRecord.objects.create(
            motorcycle=self.motorcycle,
            maintenance_type=MaintenanceType.OIL_CHANGE,
            date=date(2026, 5, 2),
            odometer_km=10000,
            cost=Decimal("120.50"),
            workshop="Oficina A",
            interval_km=3000,
            interval_days=180,
            description="Troca preventiva",
        )
        MaintenanceRecord.objects.create(
            motorcycle=self.motorcycle,
            maintenance_type=MaintenanceType.REVIEW,
            date=date(2025, 12, 20),
            odometer_km=8000,
            cost=Decimal("80.00"),
        )

        response = build_maintenance_export(
            user=self.user,
            start=date(2026, 1, 1),
            end=date(2026, 12, 31),
            fmt="xlsx",
            email_to=None,
        )
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))

        self.assertEqual(response["Content-Disposition"], 'attachment; filename="manutencoes.xlsx"')
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][1], "Moto Export")
        self.assertEqual(rows[1][2], "Troca de óleo")
        self.assertEqual(rows[1][4], 120.5)
        self.assertEqual(_amount(None), None)
        self.assertEqual(_amount(Decimal("1.23")), Decimal("1.23"))

    def test_reminders_export_respects_created_at_filter_and_email(self):
        inside = Reminder.objects.create(
            motorcycle=self.motorcycle,
            title="Troca de óleo",
            trigger_type=TriggerType.BY_INTERVAL,
            trigger_value_km=3000,
            trigger_value_days=180,
            reference_km=10000,
            reference_date=date(2026, 4, 1),
            is_active=True,
        )
        outside = Reminder.objects.create(
            motorcycle=self.motorcycle,
            title="Arquivado",
            trigger_type=TriggerType.BY_KM,
            trigger_value_km=1000,
            reference_km=5000,
            is_active=False,
        )
        Reminder.objects.filter(pk=inside.pk).update(created_at=timezone.make_aware(datetime(2026, 5, 2, 12, 0, 0)))
        Reminder.objects.filter(pk=outside.pk).update(
            created_at=timezone.make_aware(datetime(2025, 12, 30, 12, 0, 0))
        )

        response = build_reminders_export(
            user=self.user,
            start=date(2026, 1, 1),
            end=date(2026, 12, 31),
            fmt="csv",
            email_to="reminders@example.com",
        )
        payload = response.content.decode("utf-8-sig")

        self.assertIn("Troca de óleo", payload)
        self.assertNotIn("Arquivado", payload)
        self.assertIn("sim", payload)
        self.assertEqual(response["Content-Disposition"], 'attachment; filename="lembretes.csv"')
        self.assertEqual(mail.outbox[-1].attachments[0][0], "lembretes.csv")

    def test_tires_export_covers_tire_and_pressure_variants(self):
        TireRecord.objects.create(
            motorcycle=self.motorcycle,
            position=TirePosition.FRONT,
            brand_model="Pirelli Angel",
            installed_at=date(2026, 5, 1),
            installed_odometer_km=12000,
            cost=Decimal("500.00"),
            wear_percent=20,
            estimated_change_km=18000,
            is_active=True,
        )
        TireRecord.objects.create(
            motorcycle=self.other_motorcycle,
            position=TirePosition.REAR,
            brand_model="Outro pneu",
            installed_at=date(2026, 5, 1),
            installed_odometer_km=9000,
            cost=Decimal("450.00"),
            wear_percent=15,
            is_active=True,
        )
        TirePressureRecord.objects.create(
            motorcycle=self.motorcycle,
            date=date(2026, 5, 3),
            psi_front=29,
            psi_rear=33,
            notes="Frio",
        )

        tire_response = build_tires_export(
            user=self.user,
            start=date(2026, 1, 1),
            end=date(2026, 12, 31),
            fmt="csv",
            email_to=None,
            kind="tires",
        )
        pressure_response = build_tires_export(
            user=self.user,
            start=date(2026, 1, 1),
            end=date(2026, 12, 31),
            fmt="xlsx",
            email_to=None,
            kind="pressure",
        )

        self.assertIn("Pirelli Angel", tire_response.content.decode("utf-8-sig"))
        self.assertNotIn("Outro pneu", tire_response.content.decode("utf-8-sig"))
        self.assertEqual(tire_response["Content-Disposition"], 'attachment; filename="pneus.csv"')

        workbook = load_workbook(BytesIO(pressure_response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(pressure_response["Content-Disposition"], 'attachment; filename="calibragens.xlsx"')
        self.assertEqual(rows[1][1], datetime(2026, 5, 3, 0, 0))
        self.assertEqual(rows[1][2], 29)
        self.assertEqual(rows[1][3], 33)


class DemoBikeTests(TestCase):
    def test_create_demo_motorcycle_populates_related_records(self):
        user = get_user_model().objects.create_user(
            username="demo-user",
            email="demo@example.com",
            password="pass12345",
        )

        motorcycle = create_demo_motorcycle(user)

        self.assertEqual(motorcycle.owner, user)
        self.assertEqual(motorcycle.name, "Moto Demo")
        self.assertEqual(motorcycle.current_odometer_km, 8500)
        self.assertEqual(motorcycle.fuel_records.count(), 7)
        self.assertEqual(motorcycle.maintenance_records.count(), 3)
        self.assertEqual(motorcycle.tire_records.count(), 2)
        self.assertEqual(motorcycle.reminders.count(), 2)
        self.assertTrue(motorcycle.fuel_records.filter(tank_full=True).exists())
        self.assertTrue(motorcycle.reminders.filter(trigger_type=TriggerType.BY_INTERVAL).exists())
